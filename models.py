#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Copyright (c) 2021. Institute of Health and Medical Technology, Hefei Institutes of Physical Science, CAS
# @Time     : 2021/9/24 17:05
# @Author   : ZL.Z
# @Email    : zzl1124@mail.ustc.edu.cn
# @Reference: None
# @FileName : models.py
# @Software : Python3.6; PyCharm; Windows10 / Ubuntu 18.04.5 LTS (GNU/Linux 5.4.0-79-generic x86_64)
# @Hardware : Intel Core i7-4712MQ; NVIDIA GeForce 840M / 2*X640-G30(XEON 6258R 2.7G); 3*NVIDIA GeForce RTX3090
# @Version  : V3.0 - ZL.Z：2022/6/7 - 2022/6/12
#             1. 多分类改为多个二分类；
#             2. 特征与模型精简；
#             3. 增加回归任务。
#             V2.0 - ZL.Z：2021/12/6
#             OVR使用最终的集成模型.
#             V1.1 - ZL.Z：2021/10/23 - 2021/10/30
#             使用最终数据，并训练调整最优.
#             V1.0 - ZL.Z：2021/9/24 - 2021/9/30
#             First version.
# @License  : None
# @Brief    : 模型

from config import *
import os
import copy
import datetime
import numpy as np
import pandas as pd
import joblib
from typing import List
import seaborn as sns
from matplotlib import pyplot as plt
import scipy.stats as stats
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split, GridSearchCV, LeaveOneOut, cross_val_score, cross_val_predict
from sklearn.ensemble import RandomForestClassifier, RandomForestRegressor
from sklearn.pipeline import Pipeline
from sklearn.metrics import roc_curve, plot_confusion_matrix, roc_auc_score, mean_absolute_error
from imblearn.over_sampling import SVMSMOTE
from imblearn.metrics import sensitivity_specificity_support
from multiprocessing import cpu_count
import shap
from alibi.explainers import KernelShap, TreeShap
from openpyxl import load_workbook


class ModelBase:
	"""模型基类"""

	def __init__(self, feat_data: pd.DataFrame, model_name: str = "", shuffle=True, random_state=rs, opt_gs=False):
		"""
		初始化
		:param feat_data: 特征数据,pd.DataFrame格式
		:param model_name: 模型名称
		:param shuffle: 是否打乱数据集，默认打乱
		:param random_state: 随机种子
		:param opt_gs: 网格搜索中是否使用已找到的最优超参数
		"""
		self.reg = False
		if model_name.split('_')[-1] == '0-12':
			feat_data_clf = feat_data.replace({"Dysarthria Severity": {2: 1}})  # 分组2替换成组1
		elif model_name.split('_')[-1] == '1-2':
			feat_data_clf = feat_data[feat_data['Dysarthria Severity'].isin([1, 2])]
			feat_data_clf = feat_data_clf.replace({"Dysarthria Severity": {1: 0, 2: 1}})
		elif model_name.split('_')[-1] == '0-1':
			feat_data_clf = feat_data[feat_data['Dysarthria Severity'].isin([0, 1])]
		elif model_name.split('_')[-1] == 'reg':
			self.reg = True
			feat_data_clf = feat_data
		else:
			raise ValueError(f'{model_name}后半段分组表示错误，仅在0-12、1-2、0-1、reg中选取')
		x_data_clf, y_label = feat_data_clf.drop(columns=['Dysarthria Severity', 'm-FDA score']), \
							  feat_data_clf['Dysarthria Severity']
		x_data_reg = feat_data.dropna(subset=['m-FDA score']).drop(columns=['Dysarthria Severity', 'm-FDA score'])
		y_score = feat_data['m-FDA score'].dropna()
		self.feat_name_list = x_data_reg.columns.tolist()
		x_data_clf, y_label = x_data_clf.to_numpy(), y_label.to_numpy()
		x_data_reg, y_score = x_data_reg.to_numpy(), y_score.to_numpy()
		# 分层抽样，避免数据不平衡问题
		self.rs = random_state
		train_data_clf, test_data_clf, train_label, test_label = \
			train_test_split(x_data_clf, y_label, random_state=self.rs, test_size=0.3, shuffle=shuffle, stratify=y_label)
		train_data_reg, test_data_reg, train_score, test_score = \
			train_test_split(x_data_reg, y_score, random_state=self.rs, test_size=0.3, shuffle=shuffle)
		self.train_label, self.test_label = train_label, test_label
		self.train_score, self.test_score = train_score, test_score
		# 划分数据后再进行数据处理，避免测试集数据泄露
		ss = StandardScaler()  # 标准化特征
		pipe = Pipeline([('ss', ss)])
		self.train_data_clf = pipe.fit_transform(train_data_clf)
		self.test_data_clf = pipe.transform(test_data_clf)
		self.train_data_reg = pipe.fit_transform(train_data_reg)
		self.test_data_reg = pipe.transform(test_data_reg)
		if model_name.split('_')[-1] != '0-12':
			oversample = SVMSMOTE(random_state=self.rs, n_jobs=-1)  # 重采样，弥补不平衡数据（仅对训练集进行）
			self.train_data_clf, self.train_label = oversample.fit_resample(self.train_data_clf, self.train_label)
			# 绘制重采样平衡样本数量之后的各类分布情况
			# from collections import Counter
			# counter = Counter(self.train_label)
			# for k, v in counter.items():
			# 	per = v / len(self.train_label) * 100
			# 	print('Class=%d, n=%d (%.3f%%)' % (k, v, per))
			# plt.bar(counter.keys(), counter.values())
			# plt.show()
		self.model_clf_save_dir = os.path.join('./models', f'clf')  # 保存模型路径
		self.model_reg_save_dir = os.path.join('./models', f'reg')
		self.fig_save_dir = os.path.join('./results/model_results', f'{model_name}')  # 保存结果曲线图片路径
		if not os.path.exists(self.model_clf_save_dir):
			os.makedirs(self.model_clf_save_dir)
		if not os.path.exists(self.model_reg_save_dir):
			os.makedirs(self.model_reg_save_dir)
		if not os.path.exists(self.fig_save_dir):
			os.makedirs(self.fig_save_dir)
		self.model_name = model_name
		self.model_file_clf = os.path.join(self.model_clf_save_dir, f'clf_{self.model_name}.m')
		self.model_file_reg = os.path.join(self.model_reg_save_dir, f'reg_{self.model_name}.m')
		self.perform_comp_f = os.path.join(current_path, "results/PerformanceComparison.xlsx")
		self.wb = load_workbook(self.perform_comp_f)  # 模型性能比较的EXCEL文件
		self.sheet = self.wb['Sheet1']
		self.loo = LeaveOneOut()  # 留一验证
		self.opt_gs = opt_gs

	def model_train(self):
		"""
		模型训练
		"""
		pass

	def model_cv(self, clf_model, cv, n_jobs=-1, fit_params=None, isprint=True):
		"""
		模型交叉验证
		:param clf_model: 分类模型
		:param cv: 交叉验证方式
		:param n_jobs: 并行处理核数，默认-1全部CPU核数
		:param fit_params: fit参数
		:param isprint: 是否打印结果，默认打印
		:return: 评估测试集的ROC曲线下面积roc_auc、敏感性、特异性
		"""
		y_pred_proba = cross_val_predict(clf_model, self.train_data_clf, self.train_label,
										 method='predict_proba', cv=cv, n_jobs=n_jobs, fit_params=fit_params)
		roc_auc = roc_auc_score(self.train_label, y_pred_proba[:, 1])
		y_preds = np.argmax(y_pred_proba, axis=-1)
		sen, spec, sup = sensitivity_specificity_support(self.train_label, y_preds, average='binary')
		if isprint:
			print(f'LOOCV ROC-AUC: {roc_auc:.4f}')
			print(f'LOOCV Sensitivity (Recall): {sen:.4f}')
			print(f'LOOCV Specificity: {spec:.4f}\n')
		return roc_auc, sen, spec

	def model_evaluate(self, fig=False, sen_spe=False):
		"""
		模型评估
		:param fig: 显示测试集的ROC曲线以及预测曲线
		:param sen_spe: ROC曲线标题中是否显示敏感性和特异性值
		:return: 评估测试集的指标
		"""
		index = 0
		for column_cell in self.sheet.iter_rows():  # 遍历行
			index += 1
			if column_cell[0].value == self.model_name:  # 每行的首个为对应的Features
				break
		if not self.reg:
			if os.path.exists(self.model_file_clf):  # 存在已训练模型且设置加载，
				print("----------加载分类模型：{}----------".format(self.model_file_clf))
				clf_model = joblib.load(self.model_file_clf)  # 加载已训练模型
			else:
				print("分类模型不存在，无法评估，请先训练")
				return None
			y_preds = clf_model.predict(self.test_data_clf)
			y_pred_proba = clf_model.predict_proba(self.test_data_clf)
			roc_auc = roc_auc_score(self.test_label, y_pred_proba[:, 1])
			sen, spec, sup = sensitivity_specificity_support(self.test_label, y_preds, average='binary')
			print('Test set ROC-AUC: %.4f\nTest set Sensitivity (Recall): %.4f\nTest set Specificity: %.4f\n' %
				  (roc_auc, sen, spec))
			self.sheet['F' + str(index)] = round(roc_auc, 4)
			self.sheet['G' + str(index)] = round(sen, 4)
			self.sheet['H' + str(index)] = round(spec, 4)
			self.wb.save(self.perform_comp_f)  # 将结果写入模型性能比较的EXCEL文件
			if fig:  # 显示测试集的ROC曲线与混淆矩阵
				if sen_spe:
					title_t = f'ROC Curve of Detecting WD ({self.model_name})\n' \
					          f'Sensitivity = {sen:.2f}, Specificity = {spec:.2f}'
				else:
					title_t = f'ROC Curve of Detecting WD ({self.model_name})'
				plt.figure(figsize=(8, 6), dpi=300, tight_layout=True)
				plt.title(title_t, fontdict={'family': font_family, 'size': 16})
				plt.xlabel('False Positive Rate', fontdict={'family': font_family, 'size': 14})
				plt.ylabel('True Positive Rate', fontdict={'family': font_family, 'size': 14})
				plt.plot([0, 1], [0, 1], c='gray', lw=1.2, ls='--')
				fpr, tpr, thresholds = roc_curve(self.test_label, y_pred_proba[:, 1])
				plt.plot(fpr, tpr, label=f'ROC curve (AUC = {roc_auc:.2f})', c='r', ls='-', lw=2)
				plt.legend(loc="lower right", prop={'family': font_family, 'size': 12})
				plt.xlim(0.0, 1.0)
				plt.ylim(0.0, 1.0)
				for sp in plt.gca().spines:
					plt.gca().spines[sp].set_color('black')
					plt.gca().spines[sp].set_linewidth(1)
				plt.gca().tick_params(direction='out', color='black', length=5, width=1)
				plt.grid(False)
				fig_file = os.path.join(self.fig_save_dir, f"ROC_{self.model_name}.png")
				if not os.path.exists(os.path.dirname(fig_file)):
					os.makedirs(os.path.dirname(fig_file))
				plt.savefig(fig_file, dpi=600, bbox_inches='tight', pad_inches=0.2)
				plt.savefig(fig_file.replace('.png', '.svg'), bbox_inches='tight', pad_inches=0.2)
				plt.show()
				plt.close('all')
				# 绘制混淆矩阵
				plt.figure(figsize=(8, 6), dpi=300, tight_layout=True)
				disp = plot_confusion_matrix(clf_model, self.test_data_clf, self.test_label, normalize='true',
				                             colorbar=False, cmap='Blues')
				disp.ax_.set_title(f'Confusion Matrix of Detecting WD ({self.model_name})',
				                   fontdict={'family': font_family, 'size': 16})
				disp.ax_.set_xlabel('Predicted Class', fontdict={'family': font_family, 'size': 14})
				disp.ax_.set_ylabel('True Class', fontdict={'family': font_family, 'size': 14})
				for sp in plt.gca().spines:
					plt.gca().spines[sp].set_color('black')
					plt.gca().spines[sp].set_linewidth(1)
				plt.gca().tick_params(direction='out', color='black', length=5, width=1)
				plt.grid(False)
				save_fig_path = fig_file.replace(f'ROC_', 'confusion_mat_')
				plt.savefig(save_fig_path, dpi=600, bbox_inches='tight', pad_inches=0.2)
				plt.savefig(save_fig_path.replace('.png', '.svg'), bbox_inches='tight', pad_inches=0.2)
				plt.show()
				plt.close('all')
			return roc_auc, sen, spec
		else:
			if os.path.exists(self.model_file_reg):  # 存在已训练模型且设置加载，
				print("----------加载回归模型：{}----------".format(self.model_file_reg))
				reg_model = joblib.load(self.model_file_reg)  # 加载已训练模型
			else:
				print("回归模型不存在，无法评估，请先训练")
				return None
			y_preds = reg_model.predict(self.test_data_reg)
			mae = mean_absolute_error(self.test_score, y_preds)
			print('Test set MAE: %.4f\n' % mae)
			self.sheet['I' + str(index)] = round(mae, 4)
			self.wb.save(self.perform_comp_f)  # 将结果写入模型性能比较的EXCEL文件
			if fig:  # 显示测试集的预测误差曲线与预测/真实值比较
				plt.figure(figsize=(8, 6), dpi=300, tight_layout=True)
				plt.title(f"Frenchay Predict ({self.model_name})",
						  fontdict={'family': font_family, 'size': 16})
				plt.xlabel('Test set samples', fontdict={'family': font_family, 'size': 14})
				plt.ylabel('m-FDA score', fontdict={'family': font_family, 'size': 14})
				plt.plot(range(1, len(self.test_score) + 1), self.test_score, 'r-', lw=1, marker=".")
				plt.plot(range(1, len(self.test_score) + 1), y_preds, 'b-.', lw=1, marker="x")
				plt.xticks(range(1, len(self.test_score) + 1), fontproperties=font_family, fontsize=12)
				plt.yticks(fontproperties=font_family, fontsize=12)
				plt.xlim(0, len(self.test_score) + 1)
				plt.legend(["True", "Predict"], loc='upper left')
				fig_file = os.path.join(self.fig_save_dir, f"reg_{self.model_name}.png")
				plt.savefig(fig_file, dpi=600, bbox_inches='tight', pad_inches=0.2)
				plt.savefig(fig_file.replace('.png', '.svg'), bbox_inches='tight', pad_inches=0.2)
				plt.show()
				plt.close('all')
				plt.figure(figsize=(8, 6), dpi=300, tight_layout=True)
				# plt.title(f"Frenchay Predict ({self.model_name})",
				# 		  fontdict={'family': font_family, 'size': 16})
				plt.xlabel('True m-FDA score', fontdict={'family': font_family, 'size': 16})
				plt.ylabel('Predicted m-FDA score', fontdict={'family': font_family, 'size': 16})
				plt.plot([0, 150], [0, 150], c='red', lw=1.2, ls='--')
				plt.scatter(self.test_score, y_preds, s=50, c='b')
				plt.xticks(fontproperties=font_family, fontsize=12)
				plt.yticks(fontproperties=font_family, fontsize=12)
				plt.xlim(50, 130)
				plt.ylim(50, 130)
				r_p = stats.spearmanr(self.test_score, y_preds)  # 计算r和p值
				plt.text(53, 120, f'MAE={mae:.2f}\n$r={r_p[0]:.2f}, p={r_p[1]:.3g}$',
						 fontdict={'family': font_family, 'size': 16})
				fig_file = os.path.join(self.fig_save_dir, f"reg_scatter_{self.model_name}.png")
				plt.savefig(fig_file, dpi=600, bbox_inches='tight', pad_inches=0.2)
				plt.savefig(fig_file.replace('.png', '.tif'), dpi=600, bbox_inches='tight', pad_inches=0.2,
				            pil_kwargs={"compression": "tiff_lzw"})
				plt.savefig(fig_file.replace('.png', '.svg'), bbox_inches='tight', pad_inches=0.2)
				plt.show()
				plt.close('all')
			return mae

	def model_explain_reg(self, shap_type='kernel'):
		"""
		回归模型解释（利用SHAP）
		:param shap_type: SHAP类型，基于树的模型采用'tree',其他以及还未支持的模型用'kernel'
		"""
		assert shap_type in ['kernel', 'tree'], \
			'参数model_output输入错误（["kernel", "tree"]）'
		if os.path.exists(self.model_file_reg):  # 存在已训练模型且设置加载，
			print("----------加载回归模型：{}----------".format(self.model_file_reg))
			reg_model = joblib.load(self.model_file_reg)  # 加载已训练模型
		else:
			print("回归模型不存在，无法评估，请先训练")
			return None
		if shap_type == 'tree':
			explainer = TreeShap(reg_model, feature_names=self.feat_name_list, task='regression', seed=self.rs)
			explainer.fit()
			explanation = explainer.explain(self.test_data_reg)
			shap_values = copy.deepcopy(explanation.data['shap_values'])  # 深拷贝防止调用绘图时shap_values改变
		else:
			explainer = KernelShap(reg_model.predict, feature_names=self.feat_name_list, task='regression',
			                       seed=self.rs, distributed_opts={'n_cpus': cpu_count() // 2})  # 并行运行KernelShap
			explainer.fit(self.train_data_reg)
			explanation = explainer.explain(self.test_data_reg)
			shap_values = copy.deepcopy(explanation.data['shap_values'])  # 深拷贝防止调用绘图时shap_values改变
			# 也可以用SHAP包串行运行KernelShap，很慢
			# explainer = shap.KernelExplainer(reg_model.predict, self.train_data_reg,
			#                                  feature_names=self.feat_name_list)
			# shap_values = [copy.deepcopy(explainer.shap_values(self.test_data_reg))]
		# print(type(shap_values), len(shap_values), shap_values[0].shape, self.test_data_reg.shape, shap_values)
		# Global explanation summary plot: explain all the predictions in the test set
		plt.figure(figsize=(8, 6), dpi=300, tight_layout=True)
		shap.summary_plot(shap_values[-1], self.test_data_reg, max_display=10,
		                  feature_names=self.feat_name_list, axis_color='black', show=False, color_bar=False)
		plt.xlabel('SHAP value (impact on model output)', fontdict={'family': font_family, 'size': 14})
		plt.xticks(fontproperties=font_family, size=12)
		plt.yticks(fontproperties=font_family, size=14)
		plt.axvline(x=0, color='black', lw=1.5)
		plt.gca().spines['bottom'].set_color('black')
		plt.gca().tick_params(axis='x', direction='out', color='black', length=5, width=1)
		m = plt.cm.ScalarMappable(cmap=shap.plots.colors.red_blue)
		m.set_array(np.array([0, 1]))
		cb = plt.colorbar(m, ticks=[0, 1], aspect=1000)
		cb.ax.tick_params(labelsize=12, length=0)
		cb.set_ticklabels(['Low', 'High'])
		cb.set_label('Feature value', labelpad=0, fontdict={'family': font_family, 'size': 14})
		cb.outline.set_visible(False)
		bbox = cb.ax.get_window_extent().transformed(plt.gcf().dpi_scale_trans.inverted())
		cb.ax.set_aspect((bbox.height - 0.9) * 20)
		plt.grid(False)
		plt.tight_layout()
		save_fig_path = os.path.join(self.fig_save_dir, f"summary_plot_{self.model_name}.png")
		if not os.path.exists(os.path.dirname(save_fig_path)):
			os.makedirs(os.path.dirname(save_fig_path))
		plt.savefig(save_fig_path, dpi=600, bbox_inches='tight', pad_inches=0.2)
		plt.savefig(save_fig_path.replace('.png', '.tif'), dpi=600, bbox_inches='tight', pad_inches=0.2,
		            pil_kwargs={"compression": "tiff_lzw"})
		plt.savefig(save_fig_path.replace('.png', '.svg'), bbox_inches='tight', pad_inches=0.2)
		plt.show()
		plt.close()

	def model_explain_clf(self, shap_type='kernel'):
		"""
		分类模型解释（利用SHAP）
		:param shap_type: SHAP类型，基于树的模型采用'tree',其他以及还未支持的模型用'kernel'
		"""
		assert shap_type in ['kernel', 'tree'], \
			'参数model_output输入错误（["kernel", "tree"]）'
		if os.path.exists(self.model_file_clf):  # 存在已训练模型且设置加载，
			print("----------加载分类模型：{}----------".format(self.model_file_clf))
			clf_model = joblib.load(self.model_file_clf)  # 加载已训练模型
		else:
			print("分类模型不存在，无法评估，请先训练")
			return None
		if shap_type == 'tree':
			explainer = TreeShap(clf_model, feature_names=self.feat_name_list, seed=rs)
			explainer.fit()
			explanation = explainer.explain(self.test_data_clf)
			shap_values = copy.deepcopy(explanation.data['shap_values'])  # 深拷贝防止调用绘图时shap_values改变
		else:
			explainer = KernelShap(clf_model.predict_proba, feature_names=self.feat_name_list,
			                       seed=rs, distributed_opts={'n_cpus': cpu_count() // 2})  # 并行运行KernelShap
			explainer.fit(self.train_data_clf)
			explanation = explainer.explain(self.test_data_clf)
			shap_values = copy.deepcopy(explanation.data['shap_values'])  # 深拷贝防止调用绘图时shap_values改变
			# 也可以用SHAP包串行运行KernelShap，很慢
			# explainer = shap.KernelExplainer(clf_model.predict_proba, self.train_data_clf,
			#                                  feature_names=self.feat_name_list)
			# shap_values = copy.deepcopy(explainer.shap_values(self.test_data_clf))
		# print(type(shap_values), len(shap_values), shap_values[0].shape, self.test_data_clf.shape, shap_values)
		# 绘制类1的Global explanation summary plot
		plt.figure(figsize=(8, 6), dpi=300, tight_layout=True)
		plt.title(f'Global Explanation for Class 1 ({self.model_name})\n',
		          fontdict={'family': font_family, 'size': 16})
		shap.summary_plot(shap_values[-1], self.test_data_clf, max_display=10,
		                  feature_names=self.feat_name_list, axis_color='black', show=False, color_bar=False)
		plt.xlabel('SHAP value (impact on model output)', fontdict={'family': font_family, 'size': 14})
		plt.xticks(fontproperties=font_family, size=12)
		plt.yticks(fontproperties=font_family, size=14)
		plt.axvline(x=0, color='black', lw=1.5)
		plt.gca().spines['bottom'].set_color('black')
		plt.gca().tick_params(axis='x', direction='out', color='black', length=5, width=1)
		m = plt.cm.ScalarMappable(cmap=shap.plots.colors.red_blue)
		m.set_array(np.array([0, 1]))
		cb = plt.colorbar(m, ticks=[0, 1], aspect=1000)
		cb.ax.tick_params(labelsize=12, length=0)
		cb.set_ticklabels(['Low', 'High'])
		cb.set_label('Feature value', labelpad=0, fontdict={'family': font_family, 'size': 14})
		cb.outline.set_visible(False)
		bbox = cb.ax.get_window_extent().transformed(plt.gcf().dpi_scale_trans.inverted())
		cb.ax.set_aspect((bbox.height - 0.9) * 20)
		plt.grid(False)
		plt.tight_layout()
		save_fig_path = os.path.join(self.fig_save_dir, f"summary_plot_{self.model_name}.png")
		if not os.path.exists(os.path.dirname(save_fig_path)):
			os.makedirs(os.path.dirname(save_fig_path))
		plt.savefig(save_fig_path, dpi=600, bbox_inches='tight', pad_inches=0.2)
		plt.savefig(save_fig_path.replace('.png', '.svg'), bbox_inches='tight', pad_inches=0.2)
		plt.show()
		plt.close()


class RFModel(ModelBase):
	"""同质集成学习Bagging：使用随机森林"""

	def __init__(self, model_name='RF_reg', **kwargs):
		"""
		初始化
		:param **kwargs: SvmModel类__init__参数
		"""
		super().__init__(model_name=model_name, **kwargs)

	def model_train(self):
		"""
		模型训练
		:return: 模型与指标
		"""
		param_grid = {'n_estimators': range(100, 600, 100),  # 决策树的个数
		              'min_samples_split': range(2, 6),  # 分割内部节点所需要的最小样本数量
		              'min_samples_leaf': range(1, 5),  # 需要在叶子结点上的最小样本数量
		              }
		if self.opt_gs:
			if self.model_name.split('_')[-1] == '0-12':
				param_grid = {'min_samples_leaf': [1], 'min_samples_split': [4], 'n_estimators': [300]}
			elif self.model_name.split('_')[-1] == '1-2':
				param_grid = {'min_samples_leaf': [2], 'min_samples_split': [2], 'n_estimators': [300]}
			elif self.model_name.split('_')[-1] == '0-1':
				param_grid = {'min_samples_leaf': [1], 'min_samples_split': [3], 'n_estimators': [400]}
			else:
				param_grid = {'min_samples_leaf': [4], 'min_samples_split': [2], 'n_estimators': [200]}
		index = 0
		for column_cell in self.sheet.iter_rows():  # 遍历行
			index += 1
			if column_cell[0].value == self.model_name:
				break
		if not self.reg:  # 分类
			_model_clf = RandomForestClassifier(n_jobs=1, random_state=self.rs)
			grid_clf = GridSearchCV(_model_clf, param_grid, cv=5, scoring='roc_auc', n_jobs=-1)
			grid_clf.fit(self.train_data_clf, self.train_label)  # 训练模型
			model_clf = grid_clf.best_estimator_
			joblib.dump(model_clf, self.model_file_clf)  # 保存模型
			print("------ The best Classification params -------")
			print(grid_clf.best_params_)
			roc_auc, sen, spec = self.model_cv(model_clf, self.loo, n_jobs=-1)
			self.sheet['B' + str(index)] = f'{roc_auc:.4f}'
			self.sheet['C' + str(index)] = f'{sen:.4f}'
			self.sheet['D' + str(index)] = f'{spec:.4f}'
			self.sheet['J' + str(index)] = str(grid_clf.best_params_)
			self.wb.save(self.perform_comp_f)  # 将结果写入模型性能比较的EXCEL文件
			return model_clf, roc_auc, sen, spec
		else:  # Regression
			_model_reg = RandomForestRegressor(n_jobs=1, random_state=self.rs)
			grid_reg = GridSearchCV(_model_reg, param_grid, cv=5, scoring='neg_mean_absolute_error', n_jobs=-1)
			grid_reg.fit(self.train_data_reg, self.train_score)
			model_reg = grid_reg.best_estimator_  # 最优模型
			joblib.dump(model_reg, self.model_file_reg)  # 保存模型
			print("------ The best Regression params -------")
			print(grid_reg.best_params_)
			mae = get_cv_res(model_reg, self.train_data_reg, self.train_score,
							 "neg_mean_absolute_error", cv=self.loo)
			print(f'LOOCV MAE: {-mae[0]}±{mae[1]}')
			self.sheet['E' + str(index)] = f'{-mae[0]}±{mae[1]}'
			self.sheet['J' + str(index)] = str(grid_reg.best_params_)
			self.wb.save(self.perform_comp_f)  # 将结果写入模型性能比较的EXCEL文件
			return model_reg, -mae[0], mae[1]

	def model_explain(self):
		"""
		模型解释（利用SHAP）
		"""
		if self.model_name.split('_')[-1] == 'reg':
			self.model_explain_reg('tree')
		else:
			self.model_explain_clf('tree')


def get_cv_res(model, features, labels, score, cv=None, n_jobs=-1, fit_params=None):
	"""
	交叉验证并获取评价指标的均值、标准差
	:param model: 模型：最优参数对应的分类或回归
	:param features: 特征
	:param labels: 标签
	:param score: 评价指标
	:param cv: 交叉验证拆分策略，默认10折交叉验证
	:param n_jobs: 并行运行的作业数
	:param fit_params: fit参数
	:return: 评价指标的均值、标准差
	"""
	if cv is None:
		cv = 10
	res = cross_val_score(model, features, labels, cv=cv, scoring=score, n_jobs=n_jobs, fit_params=fit_params)
	return round(res.mean(), 4), round(res.std(), 4)


def roc_plot(feat_data: pd.DataFrame, model_name_list: List[str], grp_name=None):
	"""
	绘制全部模型的ROC曲线至同一张图上
	:param feat_data: 特征数据,pd.DataFrame格式
	:param model_name_list: 分类模型名称列表
    :param grp_name: None或dict类型，{'0':'first', '1':'second', '2':'third'}，将键替换为值
	:return: None
	"""
	color_list = sns.color_palette('muted', len(model_name_list))
	lw_list = ['-', '--', '-.', ':'] * len(model_name_list)
	plt.figure(figsize=(8, 6), dpi=300, tight_layout=True)
	for i_mn in model_name_list:
		model = RFModel(feat_data=feat_data, model_name=i_mn)
		test_data_clf, test_label = model.test_data_clf, model.test_label
		model_file_clf = os.path.join('models/clf', f'clf_{i_mn}.m')
		if os.path.exists(model_file_clf):  # 存在已训练模型且设置加载，
			print("----------加载分类模型：{}----------".format(model_file_clf))
			clf_model = joblib.load(model_file_clf)  # 加载已训练模型
		else:
			print(f"{model_file_clf}分类模型不存在，无法评估，请先训练")
			return None
		y_preds = clf_model.predict(test_data_clf)
		y_pred_proba = clf_model.predict_proba(test_data_clf)
		roc_auc = roc_auc_score(test_label, y_pred_proba[:, 1])
		sen, spec, sup = sensitivity_specificity_support(test_label, y_preds, average='binary')
		# title_t = f'ROC Curve of Detecting WD'
		# plt.title(title_t, fontdict={'family': font_family, 'size': 16})
		fpr, tpr, thresholds = roc_curve(test_label, y_pred_proba[:, 1])
		if isinstance(grp_name, dict):
			comp = grp_name[i_mn.split('_')[-1].split('-')[0]] + ' vs. ' + grp_name[i_mn.split('_')[-1].split('-')[-1]]
		else:
			comp = i_mn.split('_')[-1].replace('-', ' vs. ')
		plt.plot(fpr, tpr, label=f'ROC curve of {comp}\n(AUC = {roc_auc:.2f}, sen = {sen:.2f}, spe = {spec:.2f})',
		         c=color_list[model_name_list.index(i_mn)], ls=lw_list[model_name_list.index(i_mn)], lw=2)
	plt.xlabel('False Positive Rate', fontdict={'family': font_family, 'size': 16})
	plt.ylabel('True Positive Rate', fontdict={'family': font_family, 'size': 16})
	plt.xticks(fontproperties=font_family, fontsize=12)
	plt.yticks(fontproperties=font_family, fontsize=12)
	plt.plot([0, 1], [0, 1], c='gray', lw=1.2, ls='--')
	plt.legend(loc="lower right", prop={'family': font_family, 'size': 14})
	plt.xlim(0.0, 1.0)
	plt.ylim(0.0, 1.0)
	for sp in plt.gca().spines:
		plt.gca().spines[sp].set_color('black')
		plt.gca().spines[sp].set_linewidth(1)
	plt.gca().tick_params(direction='out', color='black', length=5, width=1)
	plt.grid(False)
	fig_file = os.path.join(res_path, "model_results/ROC_all.png")
	if not os.path.exists(os.path.dirname(fig_file)):
		os.makedirs(os.path.dirname(fig_file))
	plt.savefig(fig_file, dpi=600, bbox_inches='tight', pad_inches=0.2)
	plt.savefig(fig_file.replace('.png', '.tif'), dpi=600, bbox_inches='tight', pad_inches=0.2,
	            pil_kwargs={"compression": "tiff_lzw"})
	plt.savefig(fig_file.replace('.png', '.svg'), bbox_inches='tight', pad_inches=0.2)
	plt.show()


if __name__ == "__main__":
	start_time = datetime.datetime.now()
	print("---------- Running %s ----------" % os.path.basename(__file__))
	print("---------- Start Time: %s ----------" % start_time.strftime("%Y-%m-%d %H:%M:%S"))
	current_path = os.getcwd()
	feat_file = os.path.join(current_path, "data/features_used_model.csv")
	res_path = os.path.join(current_path, r"results")
	feat_data_all = pd.read_csv(feat_file).drop(columns=['SubjectID', 'Age', 'Gender', 'Edu', 'Group'])
	feat_data_all.columns = feat_data_all.columns.str.replace(r"^.*-|\(.*$", '', regex=True)  # 删除特征名中-之前字符和(之后字符
	feat_data_all.rename(columns={'Frenchay Score': 'm-FDA score'}, inplace=True)
	feat_data_all.fillna(feat_data_all.groupby('Dysarthria Severity').transform('mean'), inplace=True)
	model_l = ['RF_0-12', 'RF_0-1', 'RF_1-2', 'RF_reg', ]
	for task in model_l:
		print(f"------- {task} model -------\n")
		_model = RFModel(feat_data=feat_data_all, model_name=f'{task}', opt_gs=True)
		_model.model_train()
		_model.model_evaluate(fig=True)
		_model.model_explain()
	roc_plot(feat_data_all, model_l[:-1],
	         grp_name={'0': 'healthy', '1': 'mild', '2': 'moderate-severe', '12': 'dysarthria'})

	end_time = datetime.datetime.now()
	print("---------- End Time: %s ----------" % end_time.strftime("%Y-%m-%d %H:%M:%S"))
	print("---------- Time Used: %s ----------" % (end_time - start_time))
	with open(os.path.join(current_path, "results/finished.txt"), "w") as ff:
		ff.write("------------------ Started at %s -------------------\r\n" % start_time.strftime("%Y-%m-%d %H:%M:%S"))
		ff.write("------------------ Finished at %s -------------------\r\n" % end_time.strftime("%Y-%m-%d %H:%M:%S"))
		ff.write("------------------ Time Used %s -------------------\r\n" % (end_time - start_time))
